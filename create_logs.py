#!/usr/bin/env python
# -*- coding: utf-8 -*-


from multiprocessing.sharedctypes import Value
from threading import local
from numpy import str_
import openpyxl
from openpyxl.styles import Font

from PyQt5.QtCore import pyqtSignal
import re
import csv
import requests
import os
from pathlib import Path
import sys

import glob

import ui

class CreateLog():
    
    textBrowser = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super(CreateLog, self).__init__(parent)
        self.ui = ui.Ui_Form()
        self.ui.setupUi(self)

    def MakeLogFile(self,selected_mode,grinding_tools):

        modes = ['_ja','_en']

        for mode in modes:

            lane_avtives = [self.ui.rb_lane_number_1.isChecked(),self.ui.rb_lane_number_2.isChecked(),self.ui.rb_lane_number_3.isChecked()]
            lane_number = lane_avtives.index(True) + 1
            filename = self.ui.e_digital_specimen_number.text() + '_lane' + str(lane_number) + '_lab_book' + mode + '.xlsx'
            
            if(self.ui.l_output_volume.text() == "not selected"):
                output_path = self.ui.l_input_volume.text() +'/Lab_book/' + filename
            else:
                output_path = self.ui.l_output_volume.text() +'/Lab_book/' + filename

            if(selected_mode == "Geology and Paleontology"):
                #テンプレートからWBを作成
                template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_rock' + mode +'.xlsx')
            elif(selected_mode == "Biology"):
                template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_biology' + mode +'.xlsx')
            else:
                template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_other' + mode +'.xlsx')

            CreateLog.replace_strings(self,template_path,output_path,grinding_tools)
        
    def metadata(self,selected_mode,metadata_path,light_mode,temperature,voxel,meta_name,grinding_tools):
        if(selected_mode == "Geology and Paleontology"):
            #テンプレートからWBを作成
            template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_rock_metadata.xlsx')
        elif(selected_mode == "Biology"):
            template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_biology_metadata.xlsx')
        else:
            template_path = os.path.join(os.getcwd() , 'template/all_devV2' +'_other_metadata.xlsx')

        CreateLog.replace_strings(self,template_path,metadata_path,grinding_tools,light_mode,temperature,voxel,meta_name)

    
    def replace_strings(self,template_path,output_path,grinding_tools,light_mode=None,temperature=None,voxel=None,meta_name=None):
        font = Font(italic=True)

        file = os.path.join(os.getcwd() , 'template/value_list.csv')
        target_cell = []
        
        f = open(file,'r') 
        rows = csv.reader(f) 
        for row in rows:
            string = row[0].replace("\ufeff","")
            string = string.replace("\n","")
            target_cell.append(string)
        f.close()
        
        #市町村、加工具、その他のみ特殊
        if(self.ui.e_locality_2.text() == ""):
            locality = self.ui.e_locality.text()
        else:
            locality = self.ui.e_locality_2.text() +", "+ self.ui.e_locality.text()

        if(grinding_tools == "Single crystal diamond bite"):
            grinding_tools = grinding_tools
        else:
            grinding_tools = grinding_tools + ", " + self.ui.cmb_grinding_tools.currentText()

        if(self.ui.comb_details_of_preparation.currentText() == "others"):
            comb_details_of_preparation = self.ui.e_other_preparation.text()
        else:
            comb_details_of_preparation = self.ui.comb_details_of_preparation.currentText()

        if(self.ui.comb_lens_adpter.currentText() == "others"):
            comb_lens_adpter = self.ui.e_other_adpter.text()
        else:
            comb_lens_adpter = self.ui.comb_lens_adpter.currentText()

        if(self.ui.comb_interval_of_psms.currentText() == "other"):
            comb_interval_of_psms = self.ui.e_other_interval.text()
        else:
            comb_interval_of_psms = self.ui.comb_interval_of_psms.currentText()
        
        target_value = [self.ui.e_digital_specimen_number.text(),
                        self.ui.e_date_of_analys.text(),
                        self.ui.e_analyst.text(),
                        self.ui.e_affiliations.text(),
                        self.ui.e_laboratory_identificiation_number.text(),
                        self.ui.e_puglic_collection_number_before_destructive_analysis.text(),
                        self.ui.e_geological_age.text() + " (" +self.ui.e_geological_age_2.text() + ")",
                        self.ui.e_horizon.text() + " (" + self.ui.e_horizon_2.text() + ")",
                        locality,
                        self.ui.e_latitude.text(),
                        self.ui.e_longitude.text(),
                        self.ui.e_time_and_date_of_collection.text(),
                        self.ui.e_age.text(),
                        self.ui.e_sex.text(),
                        self.ui.e_strage_condition.text(),
                        self.ui.e_sample_provider.text(),
                        self.ui.e_references.text(),
                        self.ui.e_file_name.text(),
                        self.ui.e_size_of_sample.text(),
                        self.ui.e_weight_of_sample.text(),
                        self.ui.e_size_of_prepared_sample.text(),
                        self.ui.e_weight_of_prepared_sample.text(),
                        self.ui.e_shutter_speed.text(),
                        self.ui.e_iso.text(),
                        self.ui.e_b_a.text(),
                        self.ui.e_g_m.text(),
                        self.ui.e_style.text(),
                        self.ui.e_contrast.text(),
                        self.ui.e_saturation.text(),
                        self.ui.e_sharpness.text(),
                        self.ui.e_saturation_other.text(),
                        checkbox_flag(self.ui.cb_prepared_or_not.isChecked()),
                        comb_details_of_preparation,
                        self.ui.comb_aluminium_powder.currentText(),
                        ['1','2','3'][int(self.ui.bg_lane.checkedId()) * (-1) - 2],
                        ['Horizontal','Vertical'][int(self.ui.bg_camera_orientation.checkedId()) * (-1) - 2],
                        self.ui.comb_camera_number.currentText(),
                        self.ui.comb_lens.currentText(),
                        comb_lens_adpter,
                        self.ui.comb_pixel_shift_nulti_shooting.currentText(),
                        comb_interval_of_psms,
                        self.ui.comb_f_number.currentText(),
                        ['MF','AT'][int(self.ui.bg_focus_mode.checkedId()) * (-1) -2],
                        self.ui.e_taxon_lithology.text(),
                        self.ui.e_total_grinding_thickness_in_each_cycle.text(),
                        self.ui.e_total_grinding_thickness_in_each_finishing_process.text(),
                        self.ui.e_total_grinding_thickness.text(),
                        self.ui.e_depth_of_cut_in_rough_grinding.text(),
                        self.ui.e_depth_of_cut_in_finish_grinding.text(),
                        self.ui.e_rough_grinding_speed.text(),
                        self.ui.e_finish_grinding_speed.text(),
                        self.ui.e_coating_speed.text(),
                        self.ui.e_blowing_speed.text(),
                        self.ui.e_polishing.text(),
                        grinding_tools,
                        self.ui.e_major_axis.text(),
                        self.ui.e_minor_axis.text(),
                        ['Present','Absence'][int(self.ui.bg_presence_1.checkedId()) * (-1) - 2],
                        ['Present','Absence'][int(self.ui.bg_presence_of_steel_plate.checkedId()) * (-1) - 2],
                        self.ui.e_specimen_storage_area.text(),
                        ['Present','Absence'][int(self.ui.bg_presence_2.checkedId()) * (-1) - 2],
                        self.ui.e_detailed_information.toPlainText(),
                        self.ui.e_depository.text(),
                        self.ui.e_correspondence_of_datasets.toPlainText(),
                        self.ui.e_changes_in_voxel_z_width_within_a_same_dataset.toPlainText(),
                        checkbox_flag_3(self.ui.cb_presence_3.isChecked()),
                        self.ui.e_notes.toPlainText(),
                        self.ui.e_light_1.text(),
                        self.ui.e_light_2.text(),self.ui.e_light_3.text(),self.ui.e_light_4.text(),self.ui.e_light_5.text(),self.ui.e_light_6.text(),self.ui.e_light_7.text(),self.ui.e_light_8.text(),self.ui.e_light_9.text(),self.ui.e_light_10.text(),
                        self.ui.e_light_11.text(),self.ui.e_light_12.text(),self.ui.e_light_13.text(),self.ui.e_light_14.text(),self.ui.e_light_15.text(),self.ui.e_light_16.text(),self.ui.e_light_17.text(),self.ui.e_light_18.text(),self.ui.e_light_19.text(),self.ui.e_light_20.text(),
                        self.ui.e_light_21.text(),self.ui.e_light_22.text(),self.ui.e_light_23.text(),self.ui.e_light_24.text(),self.ui.e_light_25.text(),self.ui.e_light_26.text(),self.ui.e_light_27.text(),self.ui.e_light_28.text(),self.ui.e_light_29.text(),self.ui.e_light_30.text(),
                        self.ui.e_light_31.text(),self.ui.e_light_32.text(),self.ui.e_light_33.text(),self.ui.e_light_34.text(),self.ui.e_light_35.text(),self.ui.e_light_36.text(),self.ui.e_light_37.text(),self.ui.e_light_38.text(),self.ui.e_light_39.text(),self.ui.e_light_40.text(),
                        self.ui.e_light_41.text(),self.ui.e_light_42.text(),self.ui.e_light_43.text(),self.ui.e_light_44.text(),self.ui.e_light_45.text(),self.ui.e_light_46.text(),self.ui.e_light_47.text(),self.ui.e_light_48.text(),self.ui.e_light_49.text(),self.ui.e_light_50.text(),
                        self.ui.e_light_51.text(),self.ui.e_light_52.text(),self.ui.e_light_53.text(),self.ui.e_light_54.text(),self.ui.e_light_55.text(),self.ui.e_light_56.text(),self.ui.e_light_57.text(),self.ui.e_light_58.text(),self.ui.e_light_59.text(),self.ui.e_light_60.text(),
                        self.ui.e_light_61.text(),self.ui.e_light_62.text(),self.ui.e_light_63.text(),self.ui.e_light_64.text(),self.ui.e_light_65.text(),self.ui.e_light_66.text(),self.ui.e_light_67.text(),self.ui.e_light_68.text(),self.ui.e_light_69.text(),self.ui.e_light_70.text(),
                        self.ui.e_light_71.text(),self.ui.e_light_72.text(),self.ui.e_light_73.text(),self.ui.e_light_74.text(),self.ui.e_light_75.text(),self.ui.e_light_76.text(),self.ui.e_light_77.text(),self.ui.e_light_78.text(),self.ui.e_light_79.text(),self.ui.e_light_80.text(),
                        self.ui.e_light_81.text(),self.ui.e_light_82.text(),self.ui.e_light_83.text(),self.ui.e_light_84.text(),self.ui.e_light_85.text(),self.ui.e_light_86.text(),self.ui.e_light_87.text(),self.ui.e_light_88.text(),self.ui.e_light_89.text(),self.ui.e_light_90.text(),
                        self.ui.e_light_91.text(),self.ui.e_light_92.text(),self.ui.e_light_93.text(),self.ui.e_light_94.text(),self.ui.e_light_95.text(),self.ui.e_light_96.text(),self.ui.e_light_97.text(),self.ui.e_light_98.text(),self.ui.e_light_99.text(),self.ui.e_light_100.text(),
                        self.ui.e_light_101.text(),self.ui.e_light_102.text(),self.ui.e_light_103.text(),self.ui.e_light_104.text(),self.ui.e_light_105.text(),self.ui.e_light_106.text(),self.ui.e_light_107.text(),self.ui.e_light_108.text(),self.ui.e_light_109.text(),self.ui.e_light_110.text(),
                        self.ui.e_light_111.text(),self.ui.e_light_112.text(),self.ui.e_light_113.text(),self.ui.e_light_114.text(),self.ui.e_light_115.text(),self.ui.e_light_116.text(),self.ui.e_light_117.text(),self.ui.e_light_118.text(),self.ui.e_light_119.text(),self.ui.e_light_120.text(),
           
                        self.ui.e_position_no_1.text(),self.ui.e_position_no_2.text(),self.ui.e_position_no_3.text(),self.ui.e_position_no_4.text(),self.ui.e_position_no_5.text(),self.ui.e_position_no_6.text(),self.ui.e_position_no_7.text(),self.ui.e_position_no_8.text(),self.ui.e_position_no_9.text(),self.ui.e_position_no_10.text(),
                        self.ui.e_position_no_11.text(),self.ui.e_position_no_12.text(),self.ui.e_position_no_13.text(),self.ui.e_position_no_14.text(),self.ui.e_position_no_15.text(),self.ui.e_position_no_16.text(),self.ui.e_position_no_17.text(),self.ui.e_position_no_18.text(),self.ui.e_position_no_19.text(),self.ui.e_position_no_20.text(),
                        self.ui.e_position_no_21.text(),self.ui.e_position_no_22.text(),self.ui.e_position_no_23.text(),self.ui.e_position_no_24.text(),self.ui.e_position_no_25.text(),self.ui.e_position_no_26.text(),self.ui.e_position_no_27.text(),self.ui.e_position_no_28.text(),self.ui.e_position_no_29.text(),self.ui.e_position_no_30.text(),
                        self.ui.e_position_no_31.text(),self.ui.e_position_no_32.text(),self.ui.e_position_no_33.text(),self.ui.e_position_no_34.text(),self.ui.e_position_no_35.text(),self.ui.e_position_no_36.text(),self.ui.e_position_no_37.text(),self.ui.e_position_no_38.text(),self.ui.e_position_no_39.text(),self.ui.e_position_no_40.text(),
                        self.ui.e_position_no_41.text(),self.ui.e_position_no_42.text(),self.ui.e_position_no_43.text(),self.ui.e_position_no_44.text(),self.ui.e_position_no_45.text(),self.ui.e_position_no_46.text(),self.ui.e_position_no_47.text(),self.ui.e_position_no_48.text(),self.ui.e_position_no_49.text(),self.ui.e_position_no_50.text(),
                        self.ui.e_position_no_51.text(),self.ui.e_position_no_52.text(),self.ui.e_position_no_53.text(),self.ui.e_position_no_54.text(),self.ui.e_position_no_55.text(),self.ui.e_position_no_56.text(),self.ui.e_position_no_57.text(),self.ui.e_position_no_58.text(),self.ui.e_position_no_59.text(),self.ui.e_position_no_60.text(),
                        self.ui.e_position_no_61.text(),self.ui.e_position_no_62.text(),self.ui.e_position_no_63.text(),self.ui.e_position_no_64.text(),self.ui.e_position_no_65.text(),self.ui.e_position_no_66.text(),self.ui.e_position_no_67.text(),self.ui.e_position_no_68.text(),self.ui.e_position_no_69.text(),self.ui.e_position_no_70.text(),
                        self.ui.e_position_no_71.text(),self.ui.e_position_no_72.text(),self.ui.e_position_no_73.text(),self.ui.e_position_no_74.text(),self.ui.e_position_no_75.text(),self.ui.e_position_no_76.text(),self.ui.e_position_no_77.text(),self.ui.e_position_no_78.text(),self.ui.e_position_no_79.text(),self.ui.e_position_no_80.text(),
                        self.ui.e_position_no_81.text(),self.ui.e_position_no_82.text(),self.ui.e_position_no_83.text(),self.ui.e_position_no_84.text(),self.ui.e_position_no_85.text(),self.ui.e_position_no_86.text(),self.ui.e_position_no_87.text(),self.ui.e_position_no_88.text(),self.ui.e_position_no_89.text(),self.ui.e_position_no_90.text(),
                        self.ui.e_position_no_91.text(),self.ui.e_position_no_92.text(),self.ui.e_position_no_93.text(),self.ui.e_position_no_94.text(),self.ui.e_position_no_95.text(),self.ui.e_position_no_96.text(),self.ui.e_position_no_97.text(),self.ui.e_position_no_98.text(),self.ui.e_position_no_99.text(),self.ui.e_position_no_100.text(),
                        self.ui.e_position_no_101.text(),self.ui.e_position_no_102.text(),self.ui.e_position_no_103.text(),self.ui.e_position_no_104.text(),self.ui.e_position_no_105.text(),self.ui.e_position_no_106.text(),self.ui.e_position_no_107.text(),self.ui.e_position_no_108.text(),self.ui.e_position_no_109.text(),self.ui.e_position_no_110.text(),
                        self.ui.e_position_no_111.text(),self.ui.e_position_no_112.text(),self.ui.e_position_no_113.text(),self.ui.e_position_no_114.text(),self.ui.e_position_no_115.text(),self.ui.e_position_no_116.text(),self.ui.e_position_no_117.text(),self.ui.e_position_no_118.text(),self.ui.e_position_no_119.text(),self.ui.e_position_no_120.text(),
                        self.ui.e_cort_1.text(),self.ui.e_cort_2.text(),self.ui.e_cort_3.text(),self.ui.e_cort_4.text(),self.ui.e_cort_5.text(),self.ui.e_cort_6.text(),self.ui.e_cort_7.text(),self.ui.e_cort_8.text(),self.ui.e_cort_9.text(),self.ui.e_cort_10.text(),
                        self.ui.e_cort_11.text(),self.ui.e_cort_12.text(),self.ui.e_cort_13.text(),self.ui.e_cort_14.text(),self.ui.e_cort_15.text(),self.ui.e_cort_16.text(),self.ui.e_cort_17.text(),self.ui.e_cort_18.text(),self.ui.e_cort_19.text(),self.ui.e_cort_20.text(),
                        self.ui.e_cort_21.text(),self.ui.e_cort_22.text(),self.ui.e_cort_23.text(),self.ui.e_cort_24.text(),self.ui.e_cort_25.text(),self.ui.e_cort_26.text(),self.ui.e_cort_27.text(),self.ui.e_cort_28.text(),self.ui.e_cort_29.text(),self.ui.e_cort_30.text(),
                        self.ui.e_cort_31.text(),self.ui.e_cort_32.text(),self.ui.e_cort_33.text(),self.ui.e_cort_34.text(),self.ui.e_cort_35.text(),self.ui.e_cort_36.text(),self.ui.e_cort_37.text(),self.ui.e_cort_38.text(),self.ui.e_cort_39.text(),self.ui.e_cort_40.text(),
                        self.ui.e_cort_41.text(),self.ui.e_cort_42.text(),self.ui.e_cort_43.text(),self.ui.e_cort_44.text(),self.ui.e_cort_45.text(),self.ui.e_cort_46.text(),self.ui.e_cort_47.text(),self.ui.e_cort_48.text(),self.ui.e_cort_49.text(),self.ui.e_cort_50.text(),
                        self.ui.e_cort_51.text(),self.ui.e_cort_52.text(),self.ui.e_cort_53.text(),self.ui.e_cort_54.text(),self.ui.e_cort_55.text(),self.ui.e_cort_56.text(),self.ui.e_cort_57.text(),self.ui.e_cort_58.text(),self.ui.e_cort_59.text(),self.ui.e_cort_60.text(),
                        self.ui.e_cort_61.text(),self.ui.e_cort_62.text(),self.ui.e_cort_63.text(),self.ui.e_cort_64.text(),self.ui.e_cort_65.text(),self.ui.e_cort_66.text(),self.ui.e_cort_67.text(),self.ui.e_cort_68.text(),self.ui.e_cort_69.text(),self.ui.e_cort_70.text(),
                        self.ui.e_cort_71.text(),self.ui.e_cort_72.text(),self.ui.e_cort_73.text(),self.ui.e_cort_74.text(),self.ui.e_cort_75.text(),self.ui.e_cort_76.text(),self.ui.e_cort_77.text(),self.ui.e_cort_78.text(),self.ui.e_cort_79.text(),self.ui.e_cort_80.text(),
                        self.ui.e_cort_81.text(),self.ui.e_cort_82.text(),self.ui.e_cort_83.text(),self.ui.e_cort_84.text(),self.ui.e_cort_85.text(),self.ui.e_cort_86.text(),self.ui.e_cort_87.text(),self.ui.e_cort_88.text(),self.ui.e_cort_89.text(),self.ui.e_cort_90.text(),
                        self.ui.e_cort_91.text(),self.ui.e_cort_92.text(),self.ui.e_cort_93.text(),self.ui.e_cort_94.text(),self.ui.e_cort_95.text(),self.ui.e_cort_96.text(),self.ui.e_cort_97.text(),self.ui.e_cort_98.text(),self.ui.e_cort_99.text(),self.ui.e_cort_100.text(),
                        self.ui.e_cort_101.text(),self.ui.e_cort_102.text(),self.ui.e_cort_103.text(),self.ui.e_cort_104.text(),self.ui.e_cort_105.text(),self.ui.e_cort_106.text(),self.ui.e_cort_107.text(),self.ui.e_cort_108.text(),self.ui.e_cort_109.text(),self.ui.e_cort_110.text(),
                        self.ui.e_cort_111.text(),self.ui.e_cort_112.text(),self.ui.e_cort_113.text(),self.ui.e_cort_114.text(),self.ui.e_cort_115.text(),self.ui.e_cort_116.text(),self.ui.e_cort_117.text(),self.ui.e_cort_118.text(),self.ui.e_cort_119.text(),self.ui.e_cort_120.text(),
                        self.ui.e_blow_1.text(),self.ui.e_blow_2.text(),self.ui.e_blow_3.text(),self.ui.e_blow_4.text(),self.ui.e_blow_5.text(),self.ui.e_blow_6.text(),self.ui.e_blow_7.text(),self.ui.e_blow_8.text(),self.ui.e_blow_9.text(),self.ui.e_blow_10.text(),
                        self.ui.e_blow_11.text(),self.ui.e_blow_12.text(),self.ui.e_blow_13.text(),self.ui.e_blow_14.text(),self.ui.e_blow_15.text(),self.ui.e_blow_16.text(),self.ui.e_blow_17.text(),self.ui.e_blow_18.text(),self.ui.e_blow_19.text(),self.ui.e_blow_20.text(),
                        self.ui.e_blow_21.text(),self.ui.e_blow_22.text(),self.ui.e_blow_23.text(),self.ui.e_blow_24.text(),self.ui.e_blow_25.text(),self.ui.e_blow_26.text(),self.ui.e_blow_27.text(),self.ui.e_blow_28.text(),self.ui.e_blow_29.text(),self.ui.e_blow_30.text(),
                        self.ui.e_blow_31.text(),self.ui.e_blow_32.text(),self.ui.e_blow_33.text(),self.ui.e_blow_34.text(),self.ui.e_blow_35.text(),self.ui.e_blow_36.text(),self.ui.e_blow_37.text(),self.ui.e_blow_38.text(),self.ui.e_blow_39.text(),self.ui.e_blow_40.text(),
                        self.ui.e_blow_41.text(),self.ui.e_blow_42.text(),self.ui.e_blow_43.text(),self.ui.e_blow_44.text(),self.ui.e_blow_45.text(),self.ui.e_blow_46.text(),self.ui.e_blow_47.text(),self.ui.e_blow_48.text(),self.ui.e_blow_49.text(),self.ui.e_blow_50.text(),
                        self.ui.e_blow_51.text(),self.ui.e_blow_52.text(),self.ui.e_blow_53.text(),self.ui.e_blow_54.text(),self.ui.e_blow_55.text(),self.ui.e_blow_56.text(),self.ui.e_blow_57.text(),self.ui.e_blow_58.text(),self.ui.e_blow_59.text(),self.ui.e_blow_60.text(),
                        self.ui.e_blow_61.text(),self.ui.e_blow_62.text(),self.ui.e_blow_63.text(),self.ui.e_blow_64.text(),self.ui.e_blow_65.text(),self.ui.e_blow_66.text(),self.ui.e_blow_67.text(),self.ui.e_blow_68.text(),self.ui.e_blow_69.text(),self.ui.e_blow_70.text(),
                        self.ui.e_blow_71.text(),self.ui.e_blow_72.text(),self.ui.e_blow_73.text(),self.ui.e_blow_74.text(),self.ui.e_blow_75.text(),self.ui.e_blow_76.text(),self.ui.e_blow_77.text(),self.ui.e_blow_78.text(),self.ui.e_blow_79.text(),self.ui.e_blow_80.text(),
                        self.ui.e_blow_81.text(),self.ui.e_blow_82.text(),self.ui.e_blow_83.text(),self.ui.e_blow_84.text(),self.ui.e_blow_85.text(),self.ui.e_blow_86.text(),self.ui.e_blow_87.text(),self.ui.e_blow_88.text(),self.ui.e_blow_89.text(),self.ui.e_blow_90.text(),
                        self.ui.e_blow_91.text(),self.ui.e_blow_92.text(),self.ui.e_blow_93.text(),self.ui.e_blow_94.text(),self.ui.e_blow_95.text(),self.ui.e_blow_96.text(),self.ui.e_blow_97.text(),self.ui.e_blow_98.text(),self.ui.e_blow_99.text(),self.ui.e_blow_100.text(),
                        self.ui.e_blow_101.text(),self.ui.e_blow_102.text(),self.ui.e_blow_103.text(),self.ui.e_blow_104.text(),self.ui.e_blow_105.text(),self.ui.e_blow_106.text(),self.ui.e_blow_107.text(),self.ui.e_blow_108.text(),self.ui.e_blow_109.text(),self.ui.e_blow_110.text(),
                        self.ui.e_blow_111.text(),self.ui.e_blow_112.text(),self.ui.e_blow_113.text(),self.ui.e_blow_114.text(),self.ui.e_blow_115.text(),self.ui.e_blow_116.text(),self.ui.e_blow_117.text(),self.ui.e_blow_118.text(),self.ui.e_blow_119.text(),self.ui.e_blow_120.text(),
                        self.ui.e_frequency_1.text(),self.ui.e_frequency_2.text(),self.ui.e_frequency_3.text(),self.ui.e_frequency_4.text(),self.ui.e_frequency_5.text(),self.ui.e_frequency_6.text(),self.ui.e_frequency_7.text(),self.ui.e_frequency_8.text(),self.ui.e_frequency_9.text(),self.ui.e_frequency_10.text(),
                        self.ui.e_frequency_11.text(),self.ui.e_frequency_12.text(),self.ui.e_frequency_13.text(),self.ui.e_frequency_14.text(),self.ui.e_frequency_15.text(),self.ui.e_frequency_16.text(),self.ui.e_frequency_17.text(),self.ui.e_frequency_18.text(),self.ui.e_frequency_19.text(),self.ui.e_frequency_20.text(),
                        self.ui.e_frequency_21.text(),self.ui.e_frequency_22.text(),self.ui.e_frequency_23.text(),self.ui.e_frequency_24.text(),self.ui.e_frequency_25.text(),self.ui.e_frequency_26.text(),self.ui.e_frequency_27.text(),self.ui.e_frequency_28.text(),self.ui.e_frequency_29.text(),self.ui.e_frequency_30.text(),
                        self.ui.e_frequency_31.text(),self.ui.e_frequency_32.text(),self.ui.e_frequency_33.text(),self.ui.e_frequency_34.text(),self.ui.e_frequency_35.text(),self.ui.e_frequency_36.text(),self.ui.e_frequency_37.text(),self.ui.e_frequency_38.text(),self.ui.e_frequency_39.text(),self.ui.e_frequency_40.text(),
                        self.ui.e_frequency_41.text(),self.ui.e_frequency_42.text(),self.ui.e_frequency_43.text(),self.ui.e_frequency_44.text(),self.ui.e_frequency_45.text(),self.ui.e_frequency_46.text(),self.ui.e_frequency_47.text(),self.ui.e_frequency_48.text(),self.ui.e_frequency_49.text(),self.ui.e_frequency_50.text(),
                        self.ui.e_frequency_51.text(),self.ui.e_frequency_52.text(),self.ui.e_frequency_53.text(),self.ui.e_frequency_54.text(),self.ui.e_frequency_55.text(),self.ui.e_frequency_56.text(),self.ui.e_frequency_57.text(),self.ui.e_frequency_58.text(),self.ui.e_frequency_59.text(),self.ui.e_frequency_60.text(),
                        self.ui.e_frequency_61.text(),self.ui.e_frequency_62.text(),self.ui.e_frequency_63.text(),self.ui.e_frequency_64.text(),self.ui.e_frequency_65.text(),self.ui.e_frequency_66.text(),self.ui.e_frequency_67.text(),self.ui.e_frequency_68.text(),self.ui.e_frequency_69.text(),self.ui.e_frequency_70.text(),
                        self.ui.e_frequency_71.text(),self.ui.e_frequency_72.text(),self.ui.e_frequency_73.text(),self.ui.e_frequency_74.text(),self.ui.e_frequency_75.text(),self.ui.e_frequency_76.text(),self.ui.e_frequency_77.text(),self.ui.e_frequency_78.text(),self.ui.e_frequency_79.text(),self.ui.e_frequency_80.text(),
                        self.ui.e_frequency_81.text(),self.ui.e_frequency_82.text(),self.ui.e_frequency_83.text(),self.ui.e_frequency_84.text(),self.ui.e_frequency_85.text(),self.ui.e_frequency_86.text(),self.ui.e_frequency_87.text(),self.ui.e_frequency_88.text(),self.ui.e_frequency_89.text(),self.ui.e_frequency_90.text(),
                        self.ui.e_frequency_91.text(),self.ui.e_frequency_92.text(),self.ui.e_frequency_93.text(),self.ui.e_frequency_94.text(),self.ui.e_frequency_95.text(),self.ui.e_frequency_96.text(),self.ui.e_frequency_97.text(),self.ui.e_frequency_98.text(),self.ui.e_frequency_99.text(),self.ui.e_frequency_100.text(),
                        self.ui.e_frequency_101.text(),self.ui.e_frequency_102.text(),self.ui.e_frequency_103.text(),self.ui.e_frequency_104.text(),self.ui.e_frequency_105.text(),self.ui.e_frequency_106.text(),self.ui.e_frequency_107.text(),self.ui.e_frequency_108.text(),self.ui.e_frequency_109.text(),self.ui.e_frequency_110.text(),
                        self.ui.e_frequency_111.text(),self.ui.e_frequency_112.text(),self.ui.e_frequency_113.text(),self.ui.e_frequency_114.text(),self.ui.e_frequency_115.text(),self.ui.e_frequency_116.text(),self.ui.e_frequency_117.text(),self.ui.e_frequency_118.text(),self.ui.e_frequency_119.text(),self.ui.e_frequency_120.text(),
                        self.ui.e_photographing_signals_1.text(),self.ui.e_photographing_signals_2.text(),self.ui.e_photographing_signals_3.text(),self.ui.e_photographing_signals_4.text(),self.ui.e_photographing_signals_5.text(),self.ui.e_photographing_signals_6.text(),self.ui.e_photographing_signals_7.text(),self.ui.e_photographing_signals_8.text(),self.ui.e_photographing_signals_9.text(),self.ui.e_photographing_signals_10.text(),
                        self.ui.e_photographing_signals_11.text(),self.ui.e_photographing_signals_12.text(),self.ui.e_photographing_signals_13.text(),self.ui.e_photographing_signals_14.text(),self.ui.e_photographing_signals_15.text(),self.ui.e_photographing_signals_16.text(),self.ui.e_photographing_signals_17.text(),self.ui.e_photographing_signals_18.text(),self.ui.e_photographing_signals_19.text(),self.ui.e_photographing_signals_20.text(),
                        self.ui.e_photographing_signals_21.text(),self.ui.e_photographing_signals_22.text(),self.ui.e_photographing_signals_23.text(),self.ui.e_photographing_signals_24.text(),self.ui.e_photographing_signals_25.text(),self.ui.e_photographing_signals_26.text(),self.ui.e_photographing_signals_27.text(),self.ui.e_photographing_signals_28.text(),self.ui.e_photographing_signals_29.text(),self.ui.e_photographing_signals_30.text(),
                        self.ui.e_photographing_signals_31.text(),self.ui.e_photographing_signals_32.text(),self.ui.e_photographing_signals_33.text(),self.ui.e_photographing_signals_34.text(),self.ui.e_photographing_signals_35.text(),self.ui.e_photographing_signals_36.text(),self.ui.e_photographing_signals_37.text(),self.ui.e_photographing_signals_38.text(),self.ui.e_photographing_signals_39.text(),self.ui.e_photographing_signals_40.text(),
                        self.ui.e_photographing_signals_41.text(),self.ui.e_photographing_signals_42.text(),self.ui.e_photographing_signals_43.text(),self.ui.e_photographing_signals_44.text(),self.ui.e_photographing_signals_45.text(),self.ui.e_photographing_signals_46.text(),self.ui.e_photographing_signals_47.text(),self.ui.e_photographing_signals_48.text(),self.ui.e_photographing_signals_49.text(),self.ui.e_photographing_signals_50.text(),
                        self.ui.e_photographing_signals_51.text(),self.ui.e_photographing_signals_52.text(),self.ui.e_photographing_signals_53.text(),self.ui.e_photographing_signals_54.text(),self.ui.e_photographing_signals_55.text(),self.ui.e_photographing_signals_56.text(),self.ui.e_photographing_signals_57.text(),self.ui.e_photographing_signals_58.text(),self.ui.e_photographing_signals_59.text(),self.ui.e_photographing_signals_60.text(),
                        self.ui.e_photographing_signals_61.text(),self.ui.e_photographing_signals_62.text(),self.ui.e_photographing_signals_63.text(),self.ui.e_photographing_signals_64.text(),self.ui.e_photographing_signals_65.text(),self.ui.e_photographing_signals_66.text(),self.ui.e_photographing_signals_67.text(),self.ui.e_photographing_signals_68.text(),self.ui.e_photographing_signals_69.text(),self.ui.e_photographing_signals_70.text(),
                        self.ui.e_photographing_signals_71.text(),self.ui.e_photographing_signals_72.text(),self.ui.e_photographing_signals_73.text(),self.ui.e_photographing_signals_74.text(),self.ui.e_photographing_signals_75.text(),self.ui.e_photographing_signals_76.text(),self.ui.e_photographing_signals_77.text(),self.ui.e_photographing_signals_78.text(),self.ui.e_photographing_signals_79.text(),self.ui.e_photographing_signals_80.text(),
                        self.ui.e_photographing_signals_81.text(),self.ui.e_photographing_signals_82.text(),self.ui.e_photographing_signals_83.text(),self.ui.e_photographing_signals_84.text(),self.ui.e_photographing_signals_85.text(),self.ui.e_photographing_signals_86.text(),self.ui.e_photographing_signals_87.text(),self.ui.e_photographing_signals_88.text(),self.ui.e_photographing_signals_89.text(),self.ui.e_photographing_signals_90.text(),
                        self.ui.e_photographing_signals_91.text(),self.ui.e_photographing_signals_92.text(),self.ui.e_photographing_signals_93.text(),self.ui.e_photographing_signals_94.text(),self.ui.e_photographing_signals_95.text(),self.ui.e_photographing_signals_96.text(),self.ui.e_photographing_signals_97.text(),self.ui.e_photographing_signals_98.text(),self.ui.e_photographing_signals_99.text(),self.ui.e_photographing_signals_100.text(),
                        self.ui.e_photographing_signals_101.text(),self.ui.e_photographing_signals_102.text(),self.ui.e_photographing_signals_103.text(),self.ui.e_photographing_signals_104.text(),self.ui.e_photographing_signals_105.text(),self.ui.e_photographing_signals_106.text(),self.ui.e_photographing_signals_107.text(),self.ui.e_photographing_signals_108.text(),self.ui.e_photographing_signals_109.text(),self.ui.e_photographing_signals_110.text(),
                        self.ui.e_photographing_signals_111.text(),self.ui.e_photographing_signals_112.text(),self.ui.e_photographing_signals_113.text(),self.ui.e_photographing_signals_114.text(),self.ui.e_photographing_signals_115.text(),self.ui.e_photographing_signals_116.text(),self.ui.e_photographing_signals_117.text(),self.ui.e_photographing_signals_118.text(),self.ui.e_photographing_signals_119.text(),self.ui.e_photographing_signals_120.text(),
                        self.ui.e_wating_time_b_1.text(),self.ui.e_wating_time_b_2.text(),self.ui.e_wating_time_b_3.text(),self.ui.e_wating_time_b_4.text(),self.ui.e_wating_time_b_5.text(),self.ui.e_wating_time_b_6.text(),self.ui.e_wating_time_b_7.text(),self.ui.e_wating_time_b_8.text(),self.ui.e_wating_time_b_9.text(),self.ui.e_wating_time_b_10.text(),
                        self.ui.e_wating_time_b_11.text(),self.ui.e_wating_time_b_12.text(),self.ui.e_wating_time_b_13.text(),self.ui.e_wating_time_b_14.text(),self.ui.e_wating_time_b_15.text(),self.ui.e_wating_time_b_16.text(),self.ui.e_wating_time_b_17.text(),self.ui.e_wating_time_b_18.text(),self.ui.e_wating_time_b_19.text(),self.ui.e_wating_time_b_20.text(),
                        self.ui.e_wating_time_b_21.text(),self.ui.e_wating_time_b_22.text(),self.ui.e_wating_time_b_23.text(),self.ui.e_wating_time_b_24.text(),self.ui.e_wating_time_b_25.text(),self.ui.e_wating_time_b_26.text(),self.ui.e_wating_time_b_27.text(),self.ui.e_wating_time_b_28.text(),self.ui.e_wating_time_b_29.text(),self.ui.e_wating_time_b_30.text(),
                        self.ui.e_wating_time_b_31.text(),self.ui.e_wating_time_b_32.text(),self.ui.e_wating_time_b_33.text(),self.ui.e_wating_time_b_34.text(),self.ui.e_wating_time_b_35.text(),self.ui.e_wating_time_b_36.text(),self.ui.e_wating_time_b_37.text(),self.ui.e_wating_time_b_38.text(),self.ui.e_wating_time_b_39.text(),self.ui.e_wating_time_b_40.text(),
                        self.ui.e_wating_time_b_41.text(),self.ui.e_wating_time_b_42.text(),self.ui.e_wating_time_b_43.text(),self.ui.e_wating_time_b_44.text(),self.ui.e_wating_time_b_45.text(),self.ui.e_wating_time_b_46.text(),self.ui.e_wating_time_b_47.text(),self.ui.e_wating_time_b_48.text(),self.ui.e_wating_time_b_49.text(),self.ui.e_wating_time_b_50.text(),
                        self.ui.e_wating_time_b_51.text(),self.ui.e_wating_time_b_52.text(),self.ui.e_wating_time_b_53.text(),self.ui.e_wating_time_b_54.text(),self.ui.e_wating_time_b_55.text(),self.ui.e_wating_time_b_56.text(),self.ui.e_wating_time_b_57.text(),self.ui.e_wating_time_b_58.text(),self.ui.e_wating_time_b_59.text(),self.ui.e_wating_time_b_60.text(),
                        self.ui.e_wating_time_b_61.text(),self.ui.e_wating_time_b_62.text(),self.ui.e_wating_time_b_63.text(),self.ui.e_wating_time_b_64.text(),self.ui.e_wating_time_b_65.text(),self.ui.e_wating_time_b_66.text(),self.ui.e_wating_time_b_67.text(),self.ui.e_wating_time_b_68.text(),self.ui.e_wating_time_b_69.text(),self.ui.e_wating_time_b_70.text(),
                        self.ui.e_wating_time_b_71.text(),self.ui.e_wating_time_b_72.text(),self.ui.e_wating_time_b_73.text(),self.ui.e_wating_time_b_74.text(),self.ui.e_wating_time_b_75.text(),self.ui.e_wating_time_b_76.text(),self.ui.e_wating_time_b_77.text(),self.ui.e_wating_time_b_78.text(),self.ui.e_wating_time_b_79.text(),self.ui.e_wating_time_b_80.text(),
                        self.ui.e_wating_time_b_81.text(),self.ui.e_wating_time_b_82.text(),self.ui.e_wating_time_b_83.text(),self.ui.e_wating_time_b_84.text(),self.ui.e_wating_time_b_85.text(),self.ui.e_wating_time_b_86.text(),self.ui.e_wating_time_b_87.text(),self.ui.e_wating_time_b_88.text(),self.ui.e_wating_time_b_89.text(),self.ui.e_wating_time_b_90.text(),
                        self.ui.e_wating_time_b_91.text(),self.ui.e_wating_time_b_92.text(),self.ui.e_wating_time_b_93.text(),self.ui.e_wating_time_b_94.text(),self.ui.e_wating_time_b_95.text(),self.ui.e_wating_time_b_96.text(),self.ui.e_wating_time_b_97.text(),self.ui.e_wating_time_b_98.text(),self.ui.e_wating_time_b_99.text(),self.ui.e_wating_time_b_100.text(),
                        self.ui.e_wating_time_b_101.text(),self.ui.e_wating_time_b_102.text(),self.ui.e_wating_time_b_103.text(),self.ui.e_wating_time_b_104.text(),self.ui.e_wating_time_b_105.text(),self.ui.e_wating_time_b_106.text(),self.ui.e_wating_time_b_107.text(),self.ui.e_wating_time_b_108.text(),self.ui.e_wating_time_b_109.text(),self.ui.e_wating_time_b_110.text(),
                        self.ui.e_wating_time_b_111.text(),self.ui.e_wating_time_b_112.text(),self.ui.e_wating_time_b_113.text(),self.ui.e_wating_time_b_114.text(),self.ui.e_wating_time_b_115.text(),self.ui.e_wating_time_b_116.text(),self.ui.e_wating_time_b_117.text(),self.ui.e_wating_time_b_118.text(),self.ui.e_wating_time_b_119.text(),self.ui.e_wating_time_b_120.text(),
                        self.ui.e_wating_time_a_1.text(),self.ui.e_wating_time_a_2.text(),self.ui.e_wating_time_a_3.text(),self.ui.e_wating_time_a_4.text(),self.ui.e_wating_time_a_5.text(),self.ui.e_wating_time_a_6.text(),self.ui.e_wating_time_a_7.text(),self.ui.e_wating_time_a_8.text(),self.ui.e_wating_time_a_9.text(),self.ui.e_wating_time_a_10.text(),
                        self.ui.e_wating_time_a_11.text(),self.ui.e_wating_time_a_12.text(),self.ui.e_wating_time_a_13.text(),self.ui.e_wating_time_a_14.text(),self.ui.e_wating_time_a_15.text(),self.ui.e_wating_time_a_16.text(),self.ui.e_wating_time_a_17.text(),self.ui.e_wating_time_a_18.text(),self.ui.e_wating_time_a_19.text(),self.ui.e_wating_time_a_20.text(),
                        self.ui.e_wating_time_a_21.text(),self.ui.e_wating_time_a_22.text(),self.ui.e_wating_time_a_23.text(),self.ui.e_wating_time_a_24.text(),self.ui.e_wating_time_a_25.text(),self.ui.e_wating_time_a_26.text(),self.ui.e_wating_time_a_27.text(),self.ui.e_wating_time_a_28.text(),self.ui.e_wating_time_a_29.text(),self.ui.e_wating_time_a_30.text(),
                        self.ui.e_wating_time_a_31.text(),self.ui.e_wating_time_a_32.text(),self.ui.e_wating_time_a_33.text(),self.ui.e_wating_time_a_34.text(),self.ui.e_wating_time_a_35.text(),self.ui.e_wating_time_a_36.text(),self.ui.e_wating_time_a_37.text(),self.ui.e_wating_time_a_38.text(),self.ui.e_wating_time_a_39.text(),self.ui.e_wating_time_a_40.text(),
                        self.ui.e_wating_time_a_41.text(),self.ui.e_wating_time_a_42.text(),self.ui.e_wating_time_a_43.text(),self.ui.e_wating_time_a_44.text(),self.ui.e_wating_time_a_45.text(),self.ui.e_wating_time_a_46.text(),self.ui.e_wating_time_a_47.text(),self.ui.e_wating_time_a_48.text(),self.ui.e_wating_time_a_49.text(),self.ui.e_wating_time_a_50.text(),
                        self.ui.e_wating_time_a_51.text(),self.ui.e_wating_time_a_52.text(),self.ui.e_wating_time_a_53.text(),self.ui.e_wating_time_a_54.text(),self.ui.e_wating_time_a_55.text(),self.ui.e_wating_time_a_56.text(),self.ui.e_wating_time_a_57.text(),self.ui.e_wating_time_a_58.text(),self.ui.e_wating_time_a_59.text(),self.ui.e_wating_time_a_60.text(),
                        self.ui.e_wating_time_a_61.text(),self.ui.e_wating_time_a_62.text(),self.ui.e_wating_time_a_63.text(),self.ui.e_wating_time_a_64.text(),self.ui.e_wating_time_a_65.text(),self.ui.e_wating_time_a_66.text(),self.ui.e_wating_time_a_67.text(),self.ui.e_wating_time_a_68.text(),self.ui.e_wating_time_a_69.text(),self.ui.e_wating_time_a_70.text(),
                        self.ui.e_wating_time_a_71.text(),self.ui.e_wating_time_a_72.text(),self.ui.e_wating_time_a_73.text(),self.ui.e_wating_time_a_74.text(),self.ui.e_wating_time_a_75.text(),self.ui.e_wating_time_a_76.text(),self.ui.e_wating_time_a_77.text(),self.ui.e_wating_time_a_78.text(),self.ui.e_wating_time_a_79.text(),self.ui.e_wating_time_a_80.text(),
                        self.ui.e_wating_time_a_81.text(),self.ui.e_wating_time_a_82.text(),self.ui.e_wating_time_a_83.text(),self.ui.e_wating_time_a_84.text(),self.ui.e_wating_time_a_85.text(),self.ui.e_wating_time_a_86.text(),self.ui.e_wating_time_a_87.text(),self.ui.e_wating_time_a_88.text(),self.ui.e_wating_time_a_89.text(),self.ui.e_wating_time_a_90.text(),
                        self.ui.e_wating_time_a_91.text(),self.ui.e_wating_time_a_92.text(),self.ui.e_wating_time_a_93.text(),self.ui.e_wating_time_a_94.text(),self.ui.e_wating_time_a_95.text(),self.ui.e_wating_time_a_96.text(),self.ui.e_wating_time_a_97.text(),self.ui.e_wating_time_a_98.text(),self.ui.e_wating_time_a_99.text(),self.ui.e_wating_time_a_100.text(),
                        self.ui.e_wating_time_a_101.text(),self.ui.e_wating_time_a_102.text(),self.ui.e_wating_time_a_103.text(),self.ui.e_wating_time_a_104.text(),self.ui.e_wating_time_a_105.text(),self.ui.e_wating_time_a_106.text(),self.ui.e_wating_time_a_107.text(),self.ui.e_wating_time_a_108.text(),self.ui.e_wating_time_a_109.text(),self.ui.e_wating_time_a_110.text(),
                        self.ui.e_wating_time_a_111.text(),self.ui.e_wating_time_a_112.text(),self.ui.e_wating_time_a_113.text(),self.ui.e_wating_time_a_114.text(),self.ui.e_wating_time_a_115.text(),self.ui.e_wating_time_a_116.text(),self.ui.e_wating_time_a_117.text(),self.ui.e_wating_time_a_118.text(),self.ui.e_wating_time_a_119.text(),self.ui.e_wating_time_a_120.text(),
           
                        self.ui.l_d_n_1.text(),self.ui.l_d_n_2.text(),self.ui.l_d_n_3.text(),self.ui.l_d_n_4.text(),self.ui.l_d_n_5.text(),self.ui.l_d_n_6.text(),self.ui.l_d_n_7.text(),self.ui.l_d_n_8.text(),self.ui.l_d_n_9.text(),self.ui.l_d_n_10.text(),
                        self.ui.l_d_n_11.text(),self.ui.l_d_n_12.text(),self.ui.l_d_n_13.text(),self.ui.l_d_n_14.text(),self.ui.l_d_n_15.text(),self.ui.l_d_n_16.text(),self.ui.l_d_n_17.text(),self.ui.l_d_n_18.text(),self.ui.l_d_n_19.text(),self.ui.l_d_n_20.text(),
                        self.ui.l_d_n_21.text(),self.ui.l_d_n_22.text(),self.ui.l_d_n_23.text(),self.ui.l_d_n_24.text(),self.ui.l_d_n_25.text(),self.ui.l_d_n_26.text(),self.ui.l_d_n_27.text(),self.ui.l_d_n_28.text(),self.ui.l_d_n_29.text(),self.ui.l_d_n_30.text(),
                        self.ui.l_d_n_31.text(),self.ui.l_d_n_32.text(),self.ui.l_d_n_33.text(),self.ui.l_d_n_34.text(),self.ui.l_d_n_35.text(),self.ui.l_d_n_36.text(),self.ui.l_d_n_37.text(),self.ui.l_d_n_38.text(),self.ui.l_d_n_39.text(),self.ui.l_d_n_40.text(),
                        self.ui.l_d_n_41.text(),self.ui.l_d_n_42.text(),self.ui.l_d_n_43.text(),self.ui.l_d_n_44.text(),self.ui.l_d_n_45.text(),self.ui.l_d_n_46.text(),self.ui.l_d_n_47.text(),self.ui.l_d_n_48.text(),self.ui.l_d_n_49.text(),self.ui.l_d_n_50.text(),
                        self.ui.l_d_n_51.text(),self.ui.l_d_n_52.text(),self.ui.l_d_n_53.text(),self.ui.l_d_n_54.text(),self.ui.l_d_n_55.text(),self.ui.l_d_n_56.text(),self.ui.l_d_n_57.text(),self.ui.l_d_n_58.text(),self.ui.l_d_n_59.text(),self.ui.l_d_n_60.text(),
                        self.ui.l_d_n_61.text(),self.ui.l_d_n_62.text(),self.ui.l_d_n_63.text(),self.ui.l_d_n_64.text(),self.ui.l_d_n_65.text(),self.ui.l_d_n_66.text(),self.ui.l_d_n_67.text(),self.ui.l_d_n_68.text(),self.ui.l_d_n_69.text(),self.ui.l_d_n_70.text(),
                        self.ui.l_d_n_71.text(),self.ui.l_d_n_72.text(),self.ui.l_d_n_73.text(),self.ui.l_d_n_74.text(),self.ui.l_d_n_75.text(),self.ui.l_d_n_76.text(),self.ui.l_d_n_77.text(),self.ui.l_d_n_78.text(),self.ui.l_d_n_79.text(),self.ui.l_d_n_80.text(),
                        self.ui.l_d_n_81.text(),self.ui.l_d_n_82.text(),self.ui.l_d_n_83.text(),self.ui.l_d_n_84.text(),self.ui.l_d_n_85.text(),self.ui.l_d_n_86.text(),self.ui.l_d_n_87.text(),self.ui.l_d_n_88.text(),self.ui.l_d_n_89.text(),self.ui.l_d_n_90.text(),
                        self.ui.l_d_n_91.text(),self.ui.l_d_n_92.text(),self.ui.l_d_n_93.text(),self.ui.l_d_n_94.text(),self.ui.l_d_n_95.text(),self.ui.l_d_n_96.text(),self.ui.l_d_n_97.text(),self.ui.l_d_n_98.text(),self.ui.l_d_n_99.text(),self.ui.l_d_n_100.text(),
                        self.ui.l_d_n_101.text(),self.ui.l_d_n_102.text(),self.ui.l_d_n_103.text(),self.ui.l_d_n_104.text(),self.ui.l_d_n_105.text(),self.ui.l_d_n_106.text(),self.ui.l_d_n_107.text(),self.ui.l_d_n_108.text(),self.ui.l_d_n_109.text(),self.ui.l_d_n_110.text(),
                        self.ui.l_d_n_111.text(),self.ui.l_d_n_112.text(),self.ui.l_d_n_113.text(),self.ui.l_d_n_114.text(),self.ui.l_d_n_115.text(),self.ui.l_d_n_116.text(),self.ui.l_d_n_117.text(),self.ui.l_d_n_118.text(),self.ui.l_d_n_119.text(),self.ui.l_d_n_120.text(),

                        self.ui.comb_optical_type_1_1.currentText(),self.ui.comb_optical_type_2_1.currentText(),self.ui.comb_optical_type_3_1.currentText(),self.ui.comb_optical_type_4_1.currentText(),self.ui.comb_optical_type_5_1.currentText(),self.ui.comb_optical_type_6_1.currentText(),self.ui.comb_optical_type_7_1.currentText(),self.ui.comb_optical_type_8_1.currentText(),
                        self.ui.comb_optical_type_1_2.currentText(),self.ui.comb_optical_type_2_2.currentText(),self.ui.comb_optical_type_3_2.currentText(),self.ui.comb_optical_type_4_2.currentText(),self.ui.comb_optical_type_5_2.currentText(),self.ui.comb_optical_type_6_2.currentText(),self.ui.comb_optical_type_7_2.currentText(),self.ui.comb_optical_type_8_2.currentText(),
                        self.ui.comb_optical_type_1_3.currentText(),self.ui.comb_optical_type_2_3.currentText(),self.ui.comb_optical_type_3_3.currentText(),self.ui.comb_optical_type_4_3.currentText(),self.ui.comb_optical_type_5_3.currentText(),self.ui.comb_optical_type_6_3.currentText(),self.ui.comb_optical_type_7_3.currentText(),self.ui.comb_optical_type_8_3.currentText(),
                        self.ui.comb_optical_type_1_4.currentText(),self.ui.comb_optical_type_2_4.currentText(),self.ui.comb_optical_type_3_4.currentText(),self.ui.comb_optical_type_4_4.currentText(),self.ui.comb_optical_type_5_4.currentText(),self.ui.comb_optical_type_6_4.currentText(),self.ui.comb_optical_type_7_4.currentText(),self.ui.comb_optical_type_8_4.currentText(),
                        self.ui.comb_individual_number_of_light_1_1.currentText(),self.ui.comb_individual_number_of_light_2_1.currentText(),self.ui.comb_individual_number_of_light_3_1.currentText(),self.ui.comb_individual_number_of_light_4_1.currentText(),self.ui.comb_individual_number_of_light_5_1.currentText(),self.ui.comb_individual_number_of_light_6_1.currentText(),self.ui.comb_individual_number_of_light_7_1.currentText(),self.ui.comb_individual_number_of_light_8_1.currentText(),
                        self.ui.comb_individual_number_of_light_1_2.currentText(),self.ui.comb_individual_number_of_light_2_2.currentText(),self.ui.comb_individual_number_of_light_3_2.currentText(),self.ui.comb_individual_number_of_light_4_2.currentText(),self.ui.comb_individual_number_of_light_5_2.currentText(),self.ui.comb_individual_number_of_light_6_2.currentText(),self.ui.comb_individual_number_of_light_7_2.currentText(),self.ui.comb_individual_number_of_light_8_2.currentText(),
                        self.ui.comb_individual_number_of_light_1_3.currentText(),self.ui.comb_individual_number_of_light_2_3.currentText(),self.ui.comb_individual_number_of_light_3_3.currentText(),self.ui.comb_individual_number_of_light_4_3.currentText(),self.ui.comb_individual_number_of_light_5_3.currentText(),self.ui.comb_individual_number_of_light_6_3.currentText(),self.ui.comb_individual_number_of_light_7_3.currentText(),self.ui.comb_individual_number_of_light_8_3.currentText(),
                        self.ui.comb_individual_number_of_light_1_4.currentText(),self.ui.comb_individual_number_of_light_2_4.currentText(),self.ui.comb_individual_number_of_light_3_4.currentText(),self.ui.comb_individual_number_of_light_4_4.currentText(),self.ui.comb_individual_number_of_light_5_4.currentText(),self.ui.comb_individual_number_of_light_6_4.currentText(),self.ui.comb_individual_number_of_light_7_4.currentText(),self.ui.comb_individual_number_of_light_8_4.currentText(),
                        self.ui.voxels[0],self.ui.voxels[1],self.ui.voxels[2],self.ui.voxels[3],self.ui.voxels[4],self.ui.voxels[5],self.ui.voxels[6],self.ui.voxels[7],self.ui.voxels[8],self.ui.voxels[9],self.ui.voxels[10],
                        self.ui.voxels[11],self.ui.voxels[12],self.ui.voxels[13],self.ui.voxels[14],self.ui.voxels[15],self.ui.voxels[16],self.ui.voxels[17],self.ui.voxels[18],self.ui.voxels[19],self.ui.voxels[20],
                        self.ui.voxels[21],self.ui.voxels[22],self.ui.voxels[23],self.ui.voxels[24],self.ui.voxels[25],self.ui.voxels[26],self.ui.voxels[27],self.ui.voxels[28],self.ui.voxels[29],self.ui.voxels[30],
                        self.ui.voxels[31],self.ui.voxels[32],self.ui.voxels[33],self.ui.voxels[34],self.ui.voxels[35],self.ui.voxels[36],self.ui.voxels[37],self.ui.voxels[38],self.ui.voxels[39],self.ui.voxels[40],
                        self.ui.voxels[41],self.ui.voxels[42],self.ui.voxels[43],self.ui.voxels[44],self.ui.voxels[45],self.ui.voxels[46],self.ui.voxels[47],self.ui.voxels[48],self.ui.voxels[49],self.ui.voxels[50],
                        self.ui.voxels[51],self.ui.voxels[52],self.ui.voxels[53],self.ui.voxels[54],self.ui.voxels[55],self.ui.voxels[56],self.ui.voxels[57],self.ui.voxels[58],self.ui.voxels[59],self.ui.voxels[60],
                        self.ui.voxels[61],self.ui.voxels[62],self.ui.voxels[63],self.ui.voxels[64],self.ui.voxels[65],self.ui.voxels[66],self.ui.voxels[67],self.ui.voxels[68],self.ui.voxels[69],self.ui.voxels[70],
                        self.ui.voxels[71],self.ui.voxels[72],self.ui.voxels[73],self.ui.voxels[74],self.ui.voxels[75],self.ui.voxels[76],self.ui.voxels[77],self.ui.voxels[78],self.ui.voxels[79],self.ui.voxels[80],
                        self.ui.voxels[81],self.ui.voxels[82],self.ui.voxels[83],self.ui.voxels[84],self.ui.voxels[85],self.ui.voxels[86],self.ui.voxels[87],self.ui.voxels[88],self.ui.voxels[89],self.ui.voxels[90],
                        self.ui.voxels[91],self.ui.voxels[92],self.ui.voxels[93],self.ui.voxels[94],self.ui.voxels[95],self.ui.voxels[96],self.ui.voxels[97],self.ui.voxels[98],self.ui.voxels[99],self.ui.voxels[100],
                        self.ui.voxels[101],self.ui.voxels[102],self.ui.voxels[103],self.ui.voxels[104],self.ui.voxels[105],self.ui.voxels[106],self.ui.voxels[107],self.ui.voxels[108],self.ui.voxels[109],self.ui.voxels[110],
                        self.ui.voxels[111],self.ui.voxels[112],self.ui.voxels[113],self.ui.voxels[114],self.ui.voxels[115],self.ui.voxels[116],self.ui.voxels[117],self.ui.voxels[118],self.ui.voxels[119],
                        self.ui.e_color_temperatuer_1_1.text(),self.ui.e_color_temperatuer_2_1.text(),self.ui.e_color_temperatuer_3_1.text(),self.ui.e_color_temperatuer_4_1.text(),self.ui.e_color_temperatuer_5_1.text(),self.ui.e_color_temperatuer_6_1.text(),self.ui.e_color_temperatuer_7_1.text(),self.ui.e_color_temperatuer_8_1.text(),
                        self.ui.e_color_temperatuer_1_2.text(),self.ui.e_color_temperatuer_2_2.text(),self.ui.e_color_temperatuer_3_2.text(),self.ui.e_color_temperatuer_4_2.text(),self.ui.e_color_temperatuer_5_2.text(),self.ui.e_color_temperatuer_6_2.text(),self.ui.e_color_temperatuer_7_2.text(),self.ui.e_color_temperatuer_8_2.text(),
                        self.ui.e_color_temperatuer_1_3.text(),self.ui.e_color_temperatuer_2_3.text(),self.ui.e_color_temperatuer_3_3.text(),self.ui.e_color_temperatuer_4_3.text(),self.ui.e_color_temperatuer_5_3.text(),self.ui.e_color_temperatuer_6_3.text(),self.ui.e_color_temperatuer_7_3.text(),self.ui.e_color_temperatuer_8_3.text(),
                        self.ui.e_color_temperatuer_1_4.text(),self.ui.e_color_temperatuer_2_4.text(),self.ui.e_color_temperatuer_3_4.text(),self.ui.e_color_temperatuer_4_4.text(),self.ui.e_color_temperatuer_5_4.text(),self.ui.e_color_temperatuer_6_4.text(),self.ui.e_color_temperatuer_7_4.text(),self.ui.e_color_temperatuer_8_4.text(),
                        
                        self.ui.e_Intensity_1_1.text(),self.ui.e_Intensity_2_1.text(),self.ui.e_Intensity_3_1.text(),self.ui.e_Intensity_4_1.text(),self.ui.e_Intensity_5_1.text(),self.ui.e_Intensity_6_1.text(),self.ui.e_Intensity_7_1.text(),self.ui.e_Intensity_8_1.text(),
                        self.ui.e_Intensity_1_2.text(),self.ui.e_Intensity_2_2.text(),self.ui.e_Intensity_3_2.text(),self.ui.e_Intensity_4_2.text(),self.ui.e_Intensity_5_2.text(),self.ui.e_Intensity_6_2.text(),self.ui.e_Intensity_7_2.text(),self.ui.e_Intensity_8_2.text(),
                        self.ui.e_Intensity_1_3.text(),self.ui.e_Intensity_2_3.text(),self.ui.e_Intensity_3_3.text(),self.ui.e_Intensity_4_3.text(),self.ui.e_Intensity_5_3.text(),self.ui.e_Intensity_6_3.text(),self.ui.e_Intensity_7_3.text(),self.ui.e_Intensity_8_3.text(),
                        self.ui.e_Intensity_1_4.text(),self.ui.e_Intensity_2_4.text(),self.ui.e_Intensity_3_4.text(),self.ui.e_Intensity_4_4.text(),self.ui.e_Intensity_5_4.text(),self.ui.e_Intensity_6_4.text(),self.ui.e_Intensity_7_4.text(),self.ui.e_Intensity_8_4.text(),
                        
                        self.ui.cb_lane_1.currentText(),self.ui.cb_lane_2.currentText(),self.ui.cb_lane_3.currentText(),self.ui.cb_lane_4.currentText(),self.ui.cb_lane_5.currentText(),self.ui.cb_lane_6.currentText(),self.ui.cb_lane_7.currentText(),self.ui.cb_lane_8.currentText(),self.ui.cb_lane_9.currentText(),self.ui.cb_lane_10.currentText(),
                        self.ui.cb_lane_11.currentText(),self.ui.cb_lane_12.currentText(),self.ui.cb_lane_13.currentText(),self.ui.cb_lane_14.currentText(),self.ui.cb_lane_15.currentText(),self.ui.cb_lane_16.currentText(),self.ui.cb_lane_17.currentText(),self.ui.cb_lane_18.currentText(),self.ui.cb_lane_19.currentText(),self.ui.cb_lane_20.currentText(),
                        self.ui.cb_lane_21.currentText(),self.ui.cb_lane_22.currentText(),self.ui.cb_lane_23.currentText(),self.ui.cb_lane_24.currentText(),self.ui.cb_lane_25.currentText(),self.ui.cb_lane_26.currentText(),self.ui.cb_lane_27.currentText(),self.ui.cb_lane_28.currentText(),self.ui.cb_lane_29.currentText(),self.ui.cb_lane_30.currentText(),
                        self.ui.cb_lane_31.currentText(),self.ui.cb_lane_32.currentText(),self.ui.cb_lane_33.currentText(),self.ui.cb_lane_34.currentText(),self.ui.cb_lane_35.currentText(),self.ui.cb_lane_36.currentText(),self.ui.cb_lane_37.currentText(),self.ui.cb_lane_38.currentText(),self.ui.cb_lane_39.currentText(),self.ui.cb_lane_40.currentText(),
                        self.ui.cb_lane_41.currentText(),self.ui.cb_lane_42.currentText(),self.ui.cb_lane_43.currentText(),self.ui.cb_lane_44.currentText(),self.ui.cb_lane_45.currentText(),self.ui.cb_lane_46.currentText(),self.ui.cb_lane_47.currentText(),self.ui.cb_lane_48.currentText(),self.ui.cb_lane_49.currentText(),self.ui.cb_lane_50.currentText(),
                        self.ui.cb_lane_51.currentText(),self.ui.cb_lane_52.currentText(),self.ui.cb_lane_53.currentText(),self.ui.cb_lane_54.currentText(),self.ui.cb_lane_55.currentText(),self.ui.cb_lane_56.currentText(),self.ui.cb_lane_57.currentText(),self.ui.cb_lane_58.currentText(),self.ui.cb_lane_59.currentText(),self.ui.cb_lane_60.currentText(),
                        self.ui.cb_lane_61.currentText(),self.ui.cb_lane_62.currentText(),self.ui.cb_lane_63.currentText(),self.ui.cb_lane_64.currentText(),self.ui.cb_lane_65.currentText(),self.ui.cb_lane_66.currentText(),self.ui.cb_lane_67.currentText(),self.ui.cb_lane_68.currentText(),self.ui.cb_lane_69.currentText(),self.ui.cb_lane_70.currentText(),
                        self.ui.cb_lane_71.currentText(),self.ui.cb_lane_72.currentText(),self.ui.cb_lane_73.currentText(),self.ui.cb_lane_74.currentText(),self.ui.cb_lane_75.currentText(),self.ui.cb_lane_76.currentText(),self.ui.cb_lane_77.currentText(),self.ui.cb_lane_78.currentText(),self.ui.cb_lane_79.currentText(),self.ui.cb_lane_80.currentText(),
                        self.ui.cb_lane_81.currentText(),self.ui.cb_lane_82.currentText(),self.ui.cb_lane_83.currentText(),self.ui.cb_lane_84.currentText(),self.ui.cb_lane_85.currentText(),self.ui.cb_lane_86.currentText(),self.ui.cb_lane_87.currentText(),self.ui.cb_lane_88.currentText(),self.ui.cb_lane_89.currentText(),self.ui.cb_lane_90.currentText(),
                        self.ui.cb_lane_91.currentText(),self.ui.cb_lane_92.currentText(),self.ui.cb_lane_93.currentText(),self.ui.cb_lane_94.currentText(),self.ui.cb_lane_95.currentText(),self.ui.cb_lane_96.currentText(),self.ui.cb_lane_97.currentText(),self.ui.cb_lane_98.currentText(),self.ui.cb_lane_99.currentText(),self.ui.cb_lane_100.currentText(),
                        self.ui.cb_lane_101.currentText(),self.ui.cb_lane_102.currentText(),self.ui.cb_lane_103.currentText(),self.ui.cb_lane_104.currentText(),self.ui.cb_lane_105.currentText(),self.ui.cb_lane_106.currentText(),self.ui.cb_lane_107.currentText(),self.ui.cb_lane_108.currentText(),self.ui.cb_lane_109.currentText(),self.ui.cb_lane_110.currentText(),
                        self.ui.cb_lane_111.currentText(),self.ui.cb_lane_112.currentText(),self.ui.cb_lane_113.currentText(),self.ui.cb_lane_114.currentText(),self.ui.cb_lane_115.currentText(),self.ui.cb_lane_116.currentText(),self.ui.cb_lane_117.currentText(),self.ui.cb_lane_118.currentText(),self.ui.cb_lane_119.currentText(),self.ui.cb_lane_120.currentText(),
                        self.ui.e_temp_all.text(),self.ui.e_pixel_width.text(),self.ui.e_fov.text(),self.ui.e_grid_A_1.text(),self.ui.e_grid_A_2.text(),self.ui.e_grid_A_3.text(),self.ui.e_grid_A_4.text(),self.ui.e_grid_A_5.text(),self.ui.e_grid_B_1.text(),self.ui.e_grid_B_2.text(),self.ui.e_grid_B_3.text(),self.ui.e_grid_B_4.text(),
                        self.ui.e_grid_B_5.text(),
                        self.ui.e_grid_C_1.text(),self.ui.e_grid_C_2.text(),self.ui.e_grid_C_3.text(),self.ui.e_grid_C_4.text(),self.ui.e_grid_C_5.text()]
            
        wb_developer = openpyxl.load_workbook(template_path)
        ws_developer = wb_developer.active

        ws_developer.title = 'sheet1'
        

        i = 0
        #リストをループ
        for list in target_cell:
            #セルをループ
            for row in ws_developer.iter_rows():
                for cell in row:
                    #セルにリストが含まれていたら
                    if cell.value == None:
                        pass
                    else:
                        if cell.value == "$filters":
                            arr = [str(i) for i in self.select_filters]
                            cell.value = '; '.join(arr)

                        elif cell.value == "$taxon_lithology" and self.ui.cb_shatai.isChecked() == True and (str(list) == cell.value):
                            ws_developer[cell.coordinate].font = font
                            cell.value = str(target_value[i])
                            
                        elif cell.value == "$camera_number":
                            cell.value = "".join(re.findall("(?<=\().+?(?=\))",self.ui.comb_camera_number.currentText())) # ['def']
                        
                        elif (light_mode != None)and ("$wave_length" == cell.value):
                            if(light_mode == "wl"):
                                cell.value = "Photographed under white light"
                            elif("uv" in light_mode):
                                uv_num = re.sub(r"\D", "", light_mode)
                                cell.value = "Photographed under UV light (" + str(uv_num) + "nm)"
                            else:
                                cell.value = ""
                        elif(temperature!= None) and ("$temperature" == cell.value):
                            cell.value = temperature
                            
                        elif(voxel!= None) and ("$voxel" == cell.value):
                            cell.value = voxel
                        
                        elif(meta_name!= None) and ("$digital_specimen_number_meta" == cell.value):
                            cell.value = meta_name
                            
                        elif str(list) == cell.value:
                            #置換
                            cell.value = str(target_value[i])

                        #英語デフォルト値専用
                        if (cell.value == "国立科学博物館（日本）") and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "National Museum of Nature and Science, Tokyo"

                        if (cell.value == "接着") and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "bonding"
                        if (cell.value == "表面の被覆") and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "coating"
                        if (cell.value == "空隙の充填") and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "filling"
                        
                        if("カップ型砥石" in str(cell.value)) and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "Grinding cup wheel, " + self.ui.cmb_grinding_tools.currentText()
                            
                        if("単結晶ダイヤモンド" in str(cell.value)) and (("_en" in output_path) or ("_meta" in output_path)):
                            cell.value = "Single crystal diamond-bite"
                            
                            
            i = i + 1
        if(("_en" in output_path) or ("_meta" in output_path)):
            for row in ws_developer.iter_rows():
                for cell in row:
                    for cell in row:
                        if cell.value == "not selected":
                                cell.value ="unselected"

        wb_developer.save(output_path)

def checkbox_flag_3(target):
    if target:
        check = "Present"
    else:
        check = "Absent"

    return check

def checkbox_flag(target):
    if target:
        check = "Prepared"
    else:
        check = "Not prepared"

    return check

def replace_logs(file_name,target_string,metadata_flag= True):
    # ブックを取得
    book = openpyxl.load_workbook(file_name)
    # シートを取得 
    sheet = book['sheet1']
    # セルを取得
    if(metadata_flag == True):
        target_cell = sheet.cell(row=1,column=1)
        sheet.cell(row=1,column=1).value = target_cell.value.replace("NNNN",target_string)
        book.save(file_name)
        
    else:#Labbook
        for i in range(500):
            target_cell = sheet.cell(row=1 + i,column=2)
            try:
                if("NNNN" in target_cell.value):
                    sheet.cell(row=1 + i,column=2).value = target_cell.value.replace("NNNN",target_string)
                    book.save(file_name)
            except:
                pass
        
        